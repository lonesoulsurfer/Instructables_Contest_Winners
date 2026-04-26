"""
Instructables Top Contest Winners Scraper  (Playwright edition)
================================================================
Scrapes the Instructables website using a real headless browser to find
the top TOP_N users with the most contest wins, broken down by prize tier,
plus rich member profile stats.

SETUP  (one time only)
-----------------------
Open a Command Prompt / PowerShell and run:

    pip install playwright openpyxl
    playwright install chromium

RUN
---
    python instructables_top_winners.py

Output files (saved next to this script):
  - instructables_winners.xlsx       Formatted Excel workbook
  - instructables_winners.csv        Raw CSV data
  - instructables_top{TOP_N}.html    Paste-ready HTML for Instructables

COLUMNS IN OUTPUT
-----------------
  Rank | Username | Total Wins | Profile Link | Followers | Member Since
  Location | Total Instructables | Total Views | First Win | Most Recent Win
  Favourite Category | Most Popular Instructable
  [Prize tier columns: Grand Prize / First Prize / Second Prize / Third Prize
   / Runner Up / Honorable Mention / any others found]
  Contests Won
"""

import time
import csv
import sys
import os
import re
from collections import defaultdict

# ── Check dependencies ────────────────────────────────────────────────────────
missing = []
try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
except ImportError:
    missing.append("playwright")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    missing.append("openpyxl")

if missing:
    print("=" * 60)
    print("ERROR: Missing required packages: " + ", ".join(missing))
    print()
    print("Please open a Command Prompt and run:")
    print("    pip install playwright openpyxl")
    print("    playwright install chromium")
    print("=" * 60)
    input("\nPress Enter to close...")
    sys.exit(1)

# ── Config ────────────────────────────────────────────────────────────────────
BASE_URL     = "https://www.instructables.com"
HEADLESS     = True    # Set False to watch the browser work
PAGE_DELAY   = 0.8     # Seconds between page loads
MAX_CONTESTS = 0       # 0 = scrape ALL; e.g. 50 for a quick test run
TOP_N        = 50      # How many top winners to show in results (e.g. 10, 50, 100)

# Known prize tier labels in preferred display order.
PRIZE_TIERS = [
    "Grand Prize",
    "First Prize",
    "Second Prize",
    "Third Prize",
    "Runner Up",
    "Honorable Mention",
]

# Keywords used to guess a contest category from its slug
CATEGORY_KEYWORDS = {
    "woodwork": "Woodworking", "wood": "Woodworking",
    "electronic": "Electronics", "arduino": "Electronics",
    "raspberry": "Electronics", "circuit": "Electronics",
    "food": "Food", "cook": "Food", "bake": "Food", "kitchen": "Food",
    "3d-print": "3D Printing", "3dprint": "3D Printing",
    "sew": "Sewing", "fabric": "Sewing", "knit": "Sewing",
    "craft": "Crafts", "paper": "Crafts", "origami": "Crafts",
    "garden": "Gardening",
    "outdoor": "Outdoors", "camping": "Outdoors",
    "science": "Science",
    "robot": "Robotics",
    "metal": "Metalworking", "weld": "Metalworking",
    "laser": "Laser Cutting", "cnc": "CNC",
    "home": "Home", "furniture": "Home",
    "toy": "Toys", "game": "Games",
    "light": "Lighting", "led": "Lighting",
    "kid": "Kids", "school": "Education",
    "costume": "Costumes", "halloween": "Costumes",
    "holiday": "Seasonal", "christmas": "Seasonal",
    "organizat": "Organisation", "storage": "Organisation",
    "bike": "Cycling",
    "repair": "Repair",
    "reuse": "Reuse / Recycle", "recycl": "Reuse / Recycle",
    "sustainab": "Sustainability", "solar": "Sustainability",
    "audio": "Audio", "music": "Audio",
    "photo": "Photography",
    "car": "Automotive", "vehicle": "Automotive",
}


def normalise_tier(raw):
    raw = raw.strip()
    mapping = {
        r"grand":                        "Grand Prize",
        r"first|1st":                    "First Prize",
        r"second|2nd":                   "Second Prize",
        r"third|3rd":                    "Third Prize",
        r"runner":                       "Runner Up",
        r"honorable|honourable|mention": "Honorable Mention",
    }
    lower = raw.lower()
    for pattern, canonical in mapping.items():
        if re.search(pattern, lower):
            return canonical
    return raw.title()


def guess_category(slug):
    s = slug.lower()
    for keyword, category in CATEGORY_KEYWORDS.items():
        if keyword in s:
            return category
    return "General"


def favourite_category(win_contests_list):
    """Return the most common category across a user's winning contests."""
    counts = defaultdict(int)
    for slug in win_contests_list:
        counts[guess_category(slug)] += 1
    if not counts:
        return ""
    return max(counts, key=counts.get)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# STEP 1  –  Collect all contest slugs from the archive
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def get_all_contest_slugs(page):
    slugs = set()
    archive_page = 1
    print("Fetching contest archive pages...")

    while True:
        url = f"{BASE_URL}/contest/archive/?page={archive_page}"
        print(f"  Archive page {archive_page} ...", end=" ", flush=True)

        try:
            page.goto(url, wait_until="domcontentloaded", timeout=30000)
            page.wait_for_selector("a[href*='/contest/']", timeout=15000)
        except PWTimeout:
            print("timed out - assuming last page.")
            break
        except Exception as e:
            print(f"error: {e}")
            break

        links = page.eval_on_selector_all(
            "a[href*='/contest/']",
            "els => els.map(e => e.getAttribute('href'))"
        )

        found_this_page = set()
        for href in links:
            if not href:
                continue
            parts = href.strip("/").split("/")
            if len(parts) == 2 and parts[0] == "contest":
                slug = parts[1]
                if slug and slug not in ("archive", ""):
                    found_this_page.add(slug)

        if not found_this_page:
            print("no contests found - reached the end.")
            break

        new = found_this_page - slugs
        slugs |= found_this_page
        print(f"found {len(found_this_page)} (+{len(new)} new, total {len(slugs)})")
        archive_page += 1
        time.sleep(PAGE_DELAY)

    return list(slugs)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# STEP 2  –  Extract winners + prize tiers for a single contest
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def get_contest_winners(page, slug):
    """Returns list of (username, prize_tier) tuples for a contest."""
    url = f"{BASE_URL}/contest/{slug}/"
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=30000)
    except Exception:
        return []

    for selector in ["text=Winners", "button:has-text('Winners')",
                     "a:has-text('Winners')", "[data-tab='winners']"]:
        try:
            el = page.query_selector(selector)
            if el:
                el.click()
                time.sleep(0.5)
                break
        except Exception:
            pass

    try:
        page.wait_for_selector("a[href^='/member/']", timeout=8000)
    except PWTimeout:
        pass

    results = page.evaluate("""
        () => {
            const out = [];
            const headingSelectors = [
                'h2','h3','h4',
                '[class*="prize"]','[class*="winner-type"]',
                '[class*="tier"]','[class*="category"]',
            ].join(',');
            const allHeadings = Array.from(document.querySelectorAll(headingSelectors));
            const prizeKW = /grand|first|1st|second|2nd|third|3rd|runner|honorable|honourable|prize|winner/i;
            const prizeHeadings = allHeadings.filter(h => prizeKW.test(h.textContent));

            if (prizeHeadings.length > 0) {
                prizeHeadings.forEach(heading => {
                    const tier = heading.textContent.trim();
                    let node = heading.nextElementSibling;
                    while (node) {
                        Array.from(node.querySelectorAll('a[href^="/member/"]')).forEach(a => {
                            const parts = a.getAttribute('href').replace(/^\\/|\\/$/g,'').split('/');
                            if (parts.length === 2 && parts[0] === 'member')
                                out.push({tier, username: parts[1]});
                        });
                        if (node.matches && node.matches(headingSelectors) &&
                            prizeKW.test(node.textContent)) break;
                        node = node.nextElementSibling;
                    }
                });
            }
            if (out.length === 0) {
                Array.from(document.querySelectorAll('a[href^="/member/"]')).forEach(a => {
                    const parts = a.getAttribute('href').replace(/^\\/|\\/$/g,'').split('/');
                    if (parts.length === 2 && parts[0] === 'member')
                        out.push({tier: 'Winner', username: parts[1]});
                });
            }
            return out;
        }
    """)

    seen = {}
    for item in results:
        username = item.get("username", "").strip()
        tier_raw = item.get("tier", "Winner").strip()
        if username and username not in seen:
            seen[username] = normalise_tier(tier_raw)

    return list(seen.items())


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# STEP 3  –  Scrape member profile stats for each top-N winner
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def get_member_profile(page, username):
    """
    Visits a member's profile page and returns a dict:
      joined, location, instructables_count, total_views,
      followers, most_popular_url
    """
    stats = {
        "joined":              "",
        "location":            "",
        "instructables_count": "",
        "total_views":         "",
        "followers":           "",
        "most_popular_url":    "",
    }

    url = f"{BASE_URL}/member/{username}/"
    try:
        # Use networkidle so JS-rendered stats have time to appear
        page.goto(url, wait_until="networkidle", timeout=40000)
        time.sleep(1.5)   # extra breathing room for React to render stats
    except Exception:
        try:
            # Fallback: domcontentloaded + longer sleep
            page.goto(url, wait_until="domcontentloaded", timeout=30000)
            time.sleep(3)
        except Exception:
            return stats

    data = page.evaluate("""
        () => {
            const first = sel => document.querySelector(sel);
            const all   = sel => Array.from(document.querySelectorAll(sel));

            // ── Helper: get numeric value from a stat block ───────────────
            // Instructables profile stats sit in elements like:
            //   <div class="...stat..."><span>1,234</span><span>Views</span></div>
            // We look for a parent whose text contains both a number AND a label.
            function findStat(labelRegex) {
                // Strategy 1: find a leaf element whose text IS just the label,
                // then grab the nearest sibling/parent number
                for (const el of all('*')) {
                    if (el.children.length === 0 && labelRegex.test(el.textContent.trim())) {
                        // Walk up to find a numeric sibling
                        const parent = el.parentElement;
                        if (parent) {
                            for (const sib of parent.children) {
                                if (sib !== el) {
                                    const m = sib.textContent.trim().match(/^([\\d,\\.]+[KkMm]?)$/);
                                    if (m) return m[1];
                                }
                            }
                            // Also check grandparent siblings
                            const gp = parent.parentElement;
                            if (gp) {
                                for (const uncle of gp.children) {
                                    if (uncle !== parent) {
                                        const m = uncle.textContent.trim().match(/^([\\d,\\.]+[KkMm]?)$/);
                                        if (m) return m[1];
                                    }
                                }
                            }
                        }
                    }
                }
                // Strategy 2: find an element whose full text matches "NUMBER LABEL"
                for (const el of all('*')) {
                    if (el.children.length <= 2) {
                        const t = el.textContent.replace(/\\s+/g,' ').trim();
                        const m = t.match(/^([\\d,\\.]+[KkMm]?)\\s+/);
                        if (m && labelRegex.test(t)) return m[1];
                    }
                }
                return '';
            }

            // ── Followers ────────────────────────────────────────────────
            let followers = findStat(/^followers?$/i) ||
                            findStat(/follower/i);

            // ── Total Views ──────────────────────────────────────────────
            let totalViews = findStat(/^views?$/i) ||
                             findStat(/total.*view|view/i);

            // ── Instructables count ───────────────────────────────────────
            let iblesCount = findStat(/^instructables?$/i) ||
                             findStat(/instructable/i);

            // ── Joined date ───────────────────────────────────────────────
            let joined = '';
            for (const el of all('*')) {
                if (el.children.length === 0) {
                    const t = el.textContent.trim();
                    // "Joined July 2020" or "Member Since July 2020"
                    let m = t.match(/(?:joined|member since)[:\\s]+(.+)/i);
                    if (m) { joined = m[1].trim(); break; }
                    // standalone date-like strings near "joined" context
                    m = t.match(/^(\\w+ \\d{1,2}(?:st|nd|rd|th)?,?\\s*\\d{4})$/i);
                    if (m && el.closest && el.closest('[class*="join"]')) {
                        joined = m[1]; break;
                    }
                }
            }

            // ── Location ─────────────────────────────────────────────────
            let location = '';
            const locSels = [
                '[class*="location"]','[class*="city"]','[class*="country"]',
                '[itemprop="addressLocality"]','[itemprop="addressCountry"]',
                '[class*="member-location"]','[class*="profile-location"]',
            ];
            for (const s of locSels) {
                const el = first(s);
                if (el && el.textContent.trim()) { location = el.textContent.trim(); break; }
            }
            // Fallback: look for a pin emoji or "Location:" label
            if (!location) {
                for (const el of all('*')) {
                    if (el.children.length === 0) {
                        const t = el.textContent.trim();
                        if (/^📍/.test(t)) { location = t.replace(/^📍\\s*/,'').trim(); break; }
                        const m = t.match(/^location[:\\s]+(.+)/i);
                        if (m) { location = m[1].trim(); break; }
                    }
                }
            }

            // ── Most popular Instructable ─────────────────────────────────
            let mostPopularUrl = '';
            const ibleLink = first('a[href*="/id/"]');
            if (ibleLink) mostPopularUrl = 'https://www.instructables.com' +
                                           ibleLink.getAttribute('href');

            return { joined, location, iblesCount, totalViews, followers, mostPopularUrl };
        }
    """)

    stats["joined"]              = data.get("joined", "")
    stats["location"]            = data.get("location", "")
    stats["instructables_count"] = data.get("iblesCount", "")
    stats["total_views"]         = data.get("totalViews", "")
    stats["followers"]           = data.get("followers", "")
    stats["most_popular_url"]    = data.get("mostPopularUrl", "")
    return stats


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# STEP 4  –  Save CSV
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def save_csv(ranked, win_contests, prize_counts, all_tiers, profiles, script_dir):
    csv_path = os.path.join(script_dir, "instructables_winners.csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            ["Rank", "Username", "Total Wins", "Profile URL",
             "Followers", "Member Since", "Location",
             "Total Instructables", "Total Views",
             "First Win", "Most Recent Win", "Favourite Category",
             "Most Popular Instructable"] +
            all_tiers +
            ["Contest Slugs"]
        )
        for rank, username, total_wins in ranked:
            p = profiles.get(username, {})
            slugs = win_contests[username]
            writer.writerow(
                [rank, username, total_wins,
                 f"{BASE_URL}/member/{username}/",
                 p.get("followers", ""),
                 p.get("joined", ""),
                 p.get("location", ""),
                 p.get("instructables_count", ""),
                 p.get("total_views", ""),
                 slugs[0]  if slugs else "",
                 slugs[-1] if slugs else "",
                 favourite_category(slugs),
                 p.get("most_popular_url", "")] +
                [prize_counts[username].get(t, 0) for t in all_tiers] +
                [" | ".join(slugs)]
            )
    print(f"  CSV   saved to: {csv_path}")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# STEP 5  –  Save formatted Excel workbook
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def save_excel(ranked, win_contests, prize_counts, all_tiers, profiles, script_dir):
    ORANGE_DARK  = "B85C00"
    ORANGE_MID   = "E07820"
    ORANGE_LIGHT = "FFF0DC"
    BLUE_DARK    = "1A4A8A"
    BLUE_MID     = "2E6EBF"
    BLUE_LIGHT   = "DCE9F8"
    GREEN_DARK   = "1A6B3A"
    GREEN_MID    = "2E9E5A"
    GREEN_LIGHT  = "D6F0E0"
    WHITE        = "FFFFFF"
    DARK_TEXT    = "1A1A1A"
    LINK_BLUE    = "1155CC"

    TIER_FILLS = ["C0392B","E67E22","F1C40F","27AE60","2980B9","8E44AD"]

    def side(style="thin", color="D0D0D0"):
        return Side(style=style, color=color)

    thin_border   = Border(left=side(), right=side(), top=side(), bottom=side())
    orange_border = Border(left=side("medium", ORANGE_DARK), right=side("medium", ORANGE_DARK),
                           top=side("medium", ORANGE_DARK),  bottom=side("medium", ORANGE_DARK))
    blue_border   = Border(left=side("medium", BLUE_DARK),   right=side("medium", BLUE_DARK),
                           top=side("medium", BLUE_DARK),    bottom=side("medium", BLUE_DARK))
    green_border  = Border(left=side("medium", GREEN_DARK),  right=side("medium", GREEN_DARK),
                           top=side("medium", GREEN_DARK),   bottom=side("medium", GREEN_DARK))

    # ── Column definitions ────────────────────────────────────────────────────
    # Each entry: (header, width, group)
    # group: "fixed" | "profile" | "calc" | "tier" | "contests"
    fixed_cols = [
        ("Rank",         7,  "fixed"),
        ("Username",     26, "fixed"),
        ("Total Wins",   12, "fixed"),
        ("Profile Link", 16, "fixed"),
    ]
    profile_cols = [
        ("Followers",            14, "profile"),
        ("Member Since",         18, "profile"),
        ("Location",             20, "profile"),
        ("Total Instructables",  20, "profile"),
        ("Total Views",          14, "profile"),
    ]
    calc_cols = [
        ("First Win",              30, "calc"),
        ("Most Recent Win",        30, "calc"),
        ("Favourite Category",     20, "calc"),
        ("Most Popular Instructable", 20, "calc"),
    ]
    tier_col_defs  = [(t, 14, "tier")     for t in all_tiers]
    contest_col    = [("Contests Won", 90, "contests")]

    all_col_defs = fixed_cols + profile_cols + calc_cols + tier_col_defs + contest_col
    n_cols = len(all_col_defs)

    def col_ltr(n):   # 1-based column index → letter(s)
        return get_column_letter(n)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Top {TOP_N} Contest Winners"

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{col_ltr(n_cols)}1")
    t = ws["A1"]
    t.value = f"Instructables Top {TOP_N} Contest Winners"
    t.font = Font(name="Arial", bold=True, size=16, color=WHITE)
    t.fill = PatternFill("solid", fgColor=ORANGE_DARK)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    # ── Subtitle ──────────────────────────────────────────────────────────────
    ws.merge_cells(f"A2:{col_ltr(n_cols)}2")
    s = ws["A2"]
    s.value = "Ranked by total competition wins across all Instructables contests"
    s.font = Font(name="Arial", italic=True, size=10, color=ORANGE_MID)
    s.fill = PatternFill("solid", fgColor="FEF4E8")
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # ── Column headers ────────────────────────────────────────────────────────
    GROUP_COLOURS = {
        "fixed":    (ORANGE_MID,  orange_border),
        "profile":  (BLUE_MID,    blue_border),
        "calc":     (GREEN_MID,   green_border),
        "tier":     (None,        orange_border),   # per-tier colour set below
        "contests": (ORANGE_DARK, orange_border),
    }

    for col_idx, (header, width, group) in enumerate(all_col_defs, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        fill_colour, border = GROUP_COLOURS[group]
        if group == "tier":
            tier_idx = col_idx - len(fixed_cols) - len(profile_cols) - len(calc_cols) - 1
            fill_colour = TIER_FILLS[tier_idx % len(TIER_FILLS)]
        cell.fill   = PatternFill("solid", fgColor=fill_colour)
        cell.border = border
        ws.column_dimensions[col_ltr(col_idx)].width = width

    ws.row_dimensions[3].height = 36

    # ── Data rows ─────────────────────────────────────────────────────────────
    for i, (rank, username, total_wins) in enumerate(ranked):
        excel_row = i + 4
        bg        = ORANGE_LIGHT if i % 2 == 0 else WHITE
        is_top3   = rank <= 3
        p         = profiles.get(username, {})
        slugs     = win_contests[username]
        profile_url = f"{BASE_URL}/member/{username}/"

        row_data = (
            [rank, username, total_wins, None] +                      # fixed (link handled below)
            [p.get("followers",""), p.get("joined",""),
             p.get("location",""), p.get("instructables_count",""),
             p.get("total_views","")] +                               # profile
            [slugs[0] if slugs else "",
             slugs[-1] if slugs else "",
             favourite_category(slugs),
             None] +                                                   # calc (popular link handled below)
            [prize_counts[username].get(t, 0) or "" for t in all_tiers] +  # tiers
            [", ".join(slugs)]                                         # contests
        )

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(excel_row, col_idx, value)
            cell.fill   = PatternFill("solid", fgColor=bg)
            cell.border = thin_border
            cell.font   = Font(name="Arial", size=10, bold=is_top3 and col_idx <= 3,
                               color=ORANGE_DARK if (is_top3 and col_idx <= 3) else DARK_TEXT)
            cell.alignment = Alignment(
                horizontal="center" if col_idx in (1, 3) else "left",
                vertical="center",
                wrap_text=(col_idx == n_cols),
                indent=1 if col_idx == 2 else 0
            )

        # Profile hyperlink (col 4)
        lc = ws.cell(excel_row, 4)
        lc.value     = "View Profile"
        lc.hyperlink = profile_url
        lc.font      = Font(name="Arial", size=10, color=LINK_BLUE, underline="single")
        lc.alignment = Alignment(horizontal="center", vertical="center")

        # Most popular instructable hyperlink
        pop_url = p.get("most_popular_url", "")
        pop_col = len(fixed_cols) + len(profile_cols) + len(calc_cols)  # last calc col
        pc = ws.cell(excel_row, pop_col)
        if pop_url:
            pc.value     = "View Instructable"
            pc.hyperlink = pop_url
            pc.font      = Font(name="Arial", size=10, color=LINK_BLUE, underline="single")
        else:
            pc.value = ""
        pc.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[excel_row].height = 48

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{col_ltr(n_cols)}{3 + len(ranked)}"

    xlsx_path = os.path.join(script_dir, "instructables_winners.xlsx")
    wb.save(xlsx_path)
    print(f"  Excel saved to: {xlsx_path}")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# STEP 6  –  Save HTML for pasting into Instructables
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def save_html(ranked, win_contests, prize_counts, all_tiers, profiles, script_dir):
    """
    Saves a clean, condensed HTML table.

    Prize tier display strategy
    ---------------------------
    The scraper collects dozens of obscure "Judge's Prize" variants (e.g. "3D Printing
    Judges' Prize", "Rainbow Judges' Prize") that are rarely awarded and create a
    massively wide, unreadable table.  Instead we:

      * Always show the six *core* tiers (Grand Prize … Judges Prize / Honorable Mention)
      * Aggregate every other tier into a "Special" column (total count of non-core wins)

    This keeps the table readable while losing no information — a tooltip / title
    attribute on each Special cell lists the breakdown.
    """
    medals = {1: "🥇", 2: "🥈", 3: "🥉"}

    # Core tiers always shown as individual columns
    CORE_TIERS = [
        "Grand Prize", "First Prize", "Second Prize",
        "Third Prize", "Runner Up", "Judges Prize",
    ]
    # Map Honorable Mention → Judges Prize bucket for display
    JUDGES_ALIASES = {"Honorable Mention", "Judges Prize", "Judges' Prize"}

    def core_count(username, tier):
        """Return count for one of the six core display tiers."""
        if tier == "Judges Prize":
            # Merge Honorable Mention and any "Judges Prize" variant into one bucket
            return sum(
                v for t, v in prize_counts[username].items()
                if t in JUDGES_ALIASES or "judge" in t.lower()
            )
        return prize_counts[username].get(tier, 0)

    def special_count_and_tip(username):
        """Return (total, tooltip_str) for all non-core tiers."""
        items = []
        total = 0
        for tier, cnt in prize_counts[username].items():
            if cnt == 0:
                continue
            is_core = (
                tier in CORE_TIERS or
                tier in JUDGES_ALIASES or
                "judge" in tier.lower()
            )
            if not is_core:
                items.append(f"{tier}: {cnt}")
                total += cnt
        tip = " | ".join(items) if items else ""
        return total, tip

    # ── Build header cells ────────────────────────────────────────────────────
    core_headers = "".join(
        f'<th style="text-align:center;min-width:72px;white-space:nowrap;">{t}</th>'
        for t in CORE_TIERS
    )
    special_header = '<th style="text-align:center;min-width:72px;" title="Sum of all other prize types">Special</th>'

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Top {TOP_N} Instructables Contest Winners</title>
<style>
  body {{ font-family: Arial, sans-serif; font-size: 13px; margin: 16px; color: #1a1a1a; }}
  h2   {{ color: #B85C00; }}
  table {{
    border-collapse: collapse;
    width: 100%;
    table-layout: fixed;
  }}
  th, td {{
    border: 1px solid #d0d0d0;
    padding: 6px 8px;
    vertical-align: middle;
    overflow: hidden;
    text-overflow: ellipsis;
    white-space: nowrap;
  }}
  thead tr {{ background-color: #E07820; color: #fff; }}
  tbody tr:nth-child(odd)  {{ background-color: #FFF0DC; }}
  tbody tr:nth-child(even) {{ background-color: #ffffff; }}
  tbody tr:hover {{ background-color: #FFE0B2; }}
  a {{ color: #1155CC; }}
  .num  {{ text-align: center; }}
  .rank {{ text-align: center; width: 48px; }}
  .member {{ width: 140px; }}
  .wins   {{ width: 64px; }}
  .fol    {{ width: 70px; }}
  .since  {{ width: 120px; }}
  .loc    {{ width: 140px; }}
  .ibs    {{ width: 80px; }}
  .views  {{ width: 88px; }}
  .cat    {{ width: 100px; }}
  .tier   {{ width: 72px; }}
  .special {{ width: 60px; }}
  .top3 td {{ font-weight: bold; }}
</style>
</head>
<body>

<h2>Top {TOP_N} Instructables Contest Winners</h2>
<p>Members ranked by total contest wins, with a breakdown by prize tier.
   Data compiled by scraping all publicly available Instructables contests.
   Hover the <em>Special</em> column for a breakdown of rare prize types.</p>

<table>
  <colgroup>
    <col class="rank">  <col class="member"> <col class="wins">
    <col class="fol">   <col class="since">  <col class="loc">
    <col class="ibs">   <col class="views">  <col class="cat">
    <!-- 6 core tier cols + 1 special col -->
    <col class="tier"><col class="tier"><col class="tier">
    <col class="tier"><col class="tier"><col class="tier">
    <col class="special">
  </colgroup>
  <thead>
    <tr>
      <th class="rank">Rank</th>
      <th class="member" style="text-align:left;">Member</th>
      <th class="num">Total Wins</th>
      <th class="num">Followers</th>
      <th class="num">Member Since</th>
      <th style="text-align:left;">Location</th>
      <th class="num">Instructables</th>
      <th class="num">Total Views</th>
      <th class="num">Fav. Category</th>
      {core_headers}
      {special_header}
    </tr>
  </thead>
  <tbody>
"""

    for i, (rank, username, total_wins) in enumerate(ranked):
        medal        = medals.get(rank, "")
        profile_url  = f"{BASE_URL}/member/{username}/"
        rank_display = f"{medal}&nbsp;{rank}" if medal else str(rank)
        row_class    = " class=\"top3\"" if rank <= 3 else ""
        p            = profiles.get(username, {})
        slugs        = win_contests[username]
        bg           = "#FFF0DC" if i % 2 == 0 else "#FFFFFF"

        core_cells = ""
        for tier in CORE_TIERS:
            cnt = core_count(username, tier)
            disp = str(cnt) if cnt else "–"
            core_cells += f'<td class="num">{disp}</td>'

        sp_cnt, sp_tip = special_count_and_tip(username)
        sp_disp  = str(sp_cnt) if sp_cnt else "–"
        sp_title = f' title="{sp_tip}"' if sp_tip else ""
        special_cell = f'<td class="num"{sp_title}>{sp_disp}</td>'

        html += f"""    <tr{row_class} style="background-color:{bg};">
      <td class="rank">{rank_display}</td>
      <td class="member"><a href="{profile_url}" target="_blank">{username}</a></td>
      <td class="num">{total_wins}</td>
      <td class="num">{p.get('followers', '')}</td>
      <td class="num">{p.get('joined', '')}</td>
      <td>{p.get('location', '')}</td>
      <td class="num">{p.get('instructables_count', '')}</td>
      <td class="num">{p.get('total_views', '')}</td>
      <td class="num">{favourite_category(slugs)}</td>
      {core_cells}
      {special_cell}
    </tr>
"""

    html += """  </tbody>
</table>
<p><em>Data based on publicly available information from instructables.com.
   &ldquo;Special&rdquo; column aggregates niche judge's prizes; hover for details.</em></p>
</body>
</html>
"""

    html_path = os.path.join(script_dir, f"instructables_top{TOP_N}.html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  HTML  saved to: {html_path}")



def main():
    print("=" * 65)
    print("  Instructables Top Contest Winners Scraper")
    print("=" * 65)

    if MAX_CONTESTS:
        print(f"  NOTE: Test mode - limiting to first {MAX_CONTESTS} contests")

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=HEADLESS)
        context = browser.new_context(user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ))
        page = context.new_page()

        # ── Phase 1: contest archive ──────────────────────────────────────
        slugs = get_all_contest_slugs(page)
        if not slugs:
            print("\nNo contest slugs found. Try HEADLESS = False to debug.")
            browser.close()
            return

        if MAX_CONTESTS:
            slugs = slugs[:MAX_CONTESTS]

        total = len(slugs)
        print(f"\nProcessing {total} contests...\n")

        win_count    = defaultdict(int)
        win_contests = defaultdict(list)
        prize_counts = defaultdict(lambda: defaultdict(int))
        seen_tiers   = []

        for i, slug in enumerate(slugs, 1):
            print(f"  [{i:>4}/{total}]  {slug:<48}", end=" ", flush=True)
            winners = get_contest_winners(page, slug)
            for username, tier in winners:
                win_count[username] += 1
                win_contests[username].append(slug)
                prize_counts[username][tier] += 1
                if tier not in seen_tiers:
                    seen_tiers.append(tier)
            print(f"-> {len(winners)} winners")
            time.sleep(PAGE_DELAY)

        if not win_count:
            print("\nNo winner data collected. Try HEADLESS = False to debug.")
            browser.close()
            return

        # ── Build top-N ranked list ───────────────────────────────────────
        ranked_pairs = sorted(win_count.items(), key=lambda x: x[1], reverse=True)[:TOP_N]
        ranked = [(rank, user, wins) for rank, (user, wins) in enumerate(ranked_pairs, 1)]

        all_tiers = [t for t in PRIZE_TIERS if t in seen_tiers]
        for t in seen_tiers:
            if t not in all_tiers:
                all_tiers.append(t)

        # ── Phase 2: member profile stats ────────────────────────────────
        print(f"\nFetching profile stats for top {TOP_N} members...\n")
        profiles = {}
        for i, (rank, username, _) in enumerate(ranked, 1):
            print(f"  [{i:>3}/{TOP_N}]  {username:<32}", end=" ", flush=True)
            profiles[username] = get_member_profile(page, username)
            print(f"followers={profiles[username]['followers'] or '?'}  "
                  f"joined={profiles[username]['joined'] or '?'}")
            time.sleep(PAGE_DELAY)

        browser.close()

    # ── Console summary ───────────────────────────────────────────────────
    print("\n" + "=" * 65)
    print(f"  TOP {TOP_N} INSTRUCTABLES CONTEST WINNERS")
    print("=" * 65)
    print(f"  {'RANK':<6} {'USERNAME':<28} {'WINS':<6} {'FOLLOWERS':<12} {'SINCE'}")
    print("  " + "-" * 62)
    for rank, user, wins in ranked:
        p = profiles.get(user, {})
        print(f"  {rank:<6} {user:<28} {wins:<6} "
              f"{p.get('followers',''):<12} {p.get('joined','')}")
    print("=" * 65)

    # ── Save all output files ─────────────────────────────────────────────
    script_dir = os.path.dirname(os.path.abspath(__file__))
    print("\nSaving output files...")
    save_csv(ranked, win_contests, prize_counts, all_tiers, profiles, script_dir)
    save_excel(ranked, win_contests, prize_counts, all_tiers, profiles, script_dir)
    save_html(ranked, win_contests, prize_counts, all_tiers, profiles, script_dir)
    print("\nAll done!")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nStopped by user.")
    except Exception as e:
        print(f"\n\nUnexpected error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        print("\n" + "-" * 65)
        input("Press Enter to close this window...")
